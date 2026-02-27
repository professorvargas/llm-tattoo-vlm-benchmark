#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pdf_per_image_all_crops.py

Gera 1 PDF por image_id contendo:
- imagem baseline
- crops black e crops white (para cada crop/segmento existente)
- máscara colorida (mask_rgb) da imagem inteira
- tabela de predições (Gemma3 / LLaMA / Qwen) com status: PERFECT / HALLUCINATION / UNKNOWN / MISS

Fonte das predições (recomendado): planilha DL_all_vlms_baseline_black_white_tables.xlsx (aba "per_crop").

Uso (exemplo):
  cd ~/projects/llm-tattoo
  pip install -U reportlab openpyxl pillow
  python3 scripts/pdf_per_image_all_crops.py \
    --splits test_open \
    --image_ids 4954893280_86e164e92f_b,10257634316_82ecfe9f0f_z,196339401_64bbc02202_b \
    --outdir pdf_per_image_demo \
    --tables_xlsx DL_all_vlms_baseline_black_white_tables.xlsx
"""
import argparse
import os
import glob
from typing import Dict, Tuple, Any, List, Optional

from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader

# -----------------------------
# Helpers: labels / status
# -----------------------------
def _to_str(x) -> str:
    return "" if x is None else str(x)

def parse_labels(s: Any) -> List[str]:
    """
    pred_labels pode vir com ',' ou ';' como separador.
    """
    if s is None:
        return []
    ss = str(s).strip()
    if ss == "" or ss.lower() == "nan":
        return []
    # normaliza separadores
    ss = ss.replace(";", ",")
    parts = [p.strip() for p in ss.split(",")]
    parts = [p for p in parts if p]
    # remove duplicados preservando ordem
    seen = set()
    out = []
    for p in parts:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out

def compute_status(gt_crop_label: str, pred: List[str]) -> str:
    """
    Classificação simples por crop:
      - MISS: pred vazio
      - PERFECT: pred == {gt}
      - HALLUCINATION: pred contém gt + extras OU não contém gt mas pred não-vazio (FP-only)
      - UNKNOWN: pred == {"unknown"} (e gt != unknown)
    """
    gt = gt_crop_label.strip()
    pset = set(pred)

    if len(pset) == 0:
        return "MISS"

    if pset == {"unknown"} and gt != "unknown":
        return "UNKNOWN"

    if gt and (pset == {gt}):
        return "PERFECT"

    if gt and (gt in pset):
        return "HALLUCINATION" if len(pset - {gt}) > 0 else "PERFECT"

    return "HALLUCINATION"

# -----------------------------
# Helpers: filesystem
# -----------------------------
def find_first_existing(paths: List[str]) -> Optional[str]:
    for p in paths:
        if p and os.path.exists(p):
            return p
    return None

def find_baseline_image(split: str, image_id: str) -> Optional[str]:
    # tenta jpg e png
    patterns = [
        f"datasets/{split}/images/{image_id}.jpg",
        f"datasets/{split}/images/{image_id}.jpeg",
        f"datasets/{split}/images/{image_id}.png",
    ]
    return find_first_existing(patterns)

def find_mask_rgb(split: str, image_id: str) -> Optional[str]:
    patterns = [
        f"datasets/{split}/mask_rgb/{image_id}_mask.jpg",
        f"datasets/{split}/mask_rgb/{image_id}_mask.png",
        f"datasets/{split}/mask_rgb/{image_id}_mask.jpeg",
    ]
    return find_first_existing(patterns)

def list_crop_files(split: str, image_id: str) -> List[str]:
    d = f"datasets/crops_gt/{split}/{image_id}"
    if not os.path.isdir(d):
        return []
    crops = sorted([os.path.basename(p) for p in glob.glob(os.path.join(d, "*.png"))])
    return crops

def path_crop_black(split: str, image_id: str, crop_file: str) -> Optional[str]:
    return find_first_existing([f"datasets/crops_gt/{split}/{image_id}/{crop_file}"])

def path_crop_white(split: str, image_id: str, crop_file: str) -> Optional[str]:
    return find_first_existing([f"datasets/crops_gt_white/{split}/{image_id}/{crop_file}"])

# -----------------------------
# Load XLSX ("per_crop") -> dict
# -----------------------------
def load_per_crop_xlsx(xlsx_path: str) -> Dict[Tuple[str, str, str, str, str], Dict[str, Any]]:
    """
    Retorna dict com chave:
      (split, variant, image_id, crop_file, model)
    e valor com campos úteis:
      pred_labels, fp_labels_not_in_gt, n_fp, gt_labels_image, crop_gt_label_from_filename, ...
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if "per_crop" not in wb.sheetnames:
        raise RuntimeError(f"A planilha '{xlsx_path}' não possui aba 'per_crop'. Abas: {wb.sheetnames}")

    ws = wb["per_crop"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {str(h).strip(): i for i, h in enumerate(header)}

    required = ["split", "variant", "image_id", "crop_file", "pred_labels", "model", "fp_labels_not_in_gt", "n_fp",
                "gt_labels_image", "crop_gt_label_from_filename"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise RuntimeError(f"Colunas ausentes em per_crop: {missing}. Colunas presentes: {list(idx.keys())}")

    data: Dict[Tuple[str, str, str, str, str], Dict[str, Any]] = {}
    for r in rows:
        split = _to_str(r[idx["split"]]).strip()
        variant = _to_str(r[idx["variant"]]).strip()
        image_id = _to_str(r[idx["image_id"]]).strip()
        crop_file = _to_str(r[idx["crop_file"]]).strip()
        model = _to_str(r[idx["model"]]).strip()

        key = (split, variant, image_id, crop_file, model)
        data[key] = {
            "pred_labels": r[idx["pred_labels"]],
            "fp_labels_not_in_gt": r[idx["fp_labels_not_in_gt"]],
            "n_fp": r[idx["n_fp"]],
            "gt_labels_image": r[idx["gt_labels_image"]],
            "crop_gt_label_from_filename": r[idx["crop_gt_label_from_filename"]],
        }
    return data

# -----------------------------
# PDF drawing
# -----------------------------
def draw_image_fit(c: canvas.Canvas, img_path: str, x: float, y: float, w: float, h: float) -> None:
    """
    Desenha imagem preservando aspect ratio dentro da caixa (x,y,w,h).
    """
    ir = ImageReader(img_path)
    iw, ih = ir.getSize()
    if iw <= 0 or ih <= 0:
        return
    scale = min(w / iw, h / ih)
    nw, nh = iw * scale, ih * scale
    ox = x + (w - nw) / 2.0
    oy = y + (h - nh) / 2.0
    c.drawImage(ir, ox, oy, width=nw, height=nh, preserveAspectRatio=True, mask='auto')

def write_wrapped(c: canvas.Canvas, text: str, x: float, y: float, max_chars: int, line_height: float) -> float:
    """
    Escreve texto quebrando por tamanho aproximado de caracteres.
    Retorna novo y.
    """
    if not text:
        return y
    words = text.split(" ")
    line = ""
    for w in words:
        if len(line) + len(w) + 1 <= max_chars:
            line = (line + " " + w).strip()
        else:
            c.drawString(x, y, line)
            y -= line_height
            line = w
    if line:
        c.drawString(x, y, line)
        y -= line_height
    return y

def make_pdf_for_image(
    out_pdf: str,
    split: str,
    image_id: str,
    per_crop: Dict[Tuple[str, str, str, str, str], Dict[str, Any]],
    models: List[str],
) -> None:
    os.makedirs(os.path.dirname(out_pdf), exist_ok=True)

    page_w, page_h = landscape(A4)
    c = canvas.Canvas(out_pdf, pagesize=(page_w, page_h))

    baseline_path = find_baseline_image(split, image_id)
    mask_path = find_mask_rgb(split, image_id)
    crops = list_crop_files(split, image_id)

    if not crops:
        crops = []  # ainda geramos PDF com 1 página (baseline + tabela baseline)

    # para obter GT (image-level) de forma estável, pegamos de qualquer linha baseline
    gt_image = None
    for m in models:
        k = (split, "baseline", image_id, "__full_image__", m)
        if k in per_crop:
            gt_image = per_crop[k].get("gt_labels_image")
            break
    gt_image_str = _to_str(gt_image).strip()

    def draw_header(title: str):
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, page_h - 40, title)
        c.setFont("Helvetica", 10)
        c.drawString(40, page_h - 60, f"split: {split}    image_id: {image_id}")
        if gt_image_str:
            c.drawString(40, page_h - 75, f"GT (image-level): {gt_image_str}")

    # Layout constants
    margin_x = 40
    top_y = page_h - 95
    img_box_w = (page_w - 2 * margin_x - 3 * 15) / 4.0
    img_box_h = 220
    gap = 15

    # Table area
    table_top = top_y - img_box_h - 20
    line_h = 13

    def draw_images_row(crop_file: Optional[str]):
        labels = ["Baseline", "Crop black", "Crop white", "Mask"]
        paths = [
            baseline_path,
            path_crop_black(split, image_id, crop_file) if crop_file else None,
            path_crop_white(split, image_id, crop_file) if crop_file else None,
            mask_path,
        ]
        x = margin_x
        y = top_y - img_box_h
        for lab, p in zip(labels, paths):
            c.setFont("Helvetica-Bold", 10)
            c.drawString(x, top_y + 2, lab)
            c.rect(x, y, img_box_w, img_box_h, stroke=1, fill=0)
            if p and os.path.exists(p):
                draw_image_fit(c, p, x + 2, y + 2, img_box_w - 4, img_box_h - 18)
            else:
                c.setFont("Helvetica", 9)
                c.drawString(x + 4, y + img_box_h / 2, "(missing)")
            x += img_box_w + gap

    def row_for(model: str, variant: str, crop_file: str) -> Dict[str, Any]:
        k = (split, variant, image_id, crop_file, model)
        return per_crop.get(k, {})

    def row_baseline(model: str) -> Dict[str, Any]:
        return row_for(model, "baseline", "__full_image__")

    def row_black(model: str, crop_file: str) -> Dict[str, Any]:
        return row_for(model, "crops_black", crop_file)

    def row_white(model: str, crop_file: str) -> Dict[str, Any]:
        return row_for(model, "crops_white", crop_file)

    def format_pred_block(gt_crop: str, d: Dict[str, Any]) -> Tuple[str, str, str, int]:
        pred = parse_labels(d.get("pred_labels"))
        # fp pode vir pronto na planilha; se estiver vazio, calculamos por pred-gt
        fp_raw = d.get("fp_labels_not_in_gt")
        fp = parse_labels(fp_raw)
        if not fp and pred:
            fp = [p for p in pred if p != gt_crop]
        status = compute_status(gt_crop, pred)
        return status, ",".join(pred), ",".join(fp), int(d.get("n_fp") or 0)

    if not crops:
        # PDF só com baseline
        draw_header("VLM per-image PDF (no crops found)")
        draw_images_row(None)
        y = table_top
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margin_x, y, "Predictions (baseline)")
        y -= 18
        c.setFont("Helvetica", 9)
        for model in models:
            base = row_baseline(model)
            status, pred_s, fp_s, nfp = format_pred_block("", base)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(margin_x, y, f"{model}: {status}")
            y -= line_h
            c.setFont("Helvetica", 9)
            y = write_wrapped(c, f"pred=[{pred_s}]", margin_x + 14, y, 120, line_h)
            y = write_wrapped(c, f"fp=[{fp_s}] n_fp={nfp}", margin_x + 14, y, 120, line_h)
            y -= 4
        c.showPage()
        c.save()
        return

    # 1 página por crop (tudo no MESMO PDF)
    for i, crop_file in enumerate(crops, start=1):
        crop_gt = os.path.splitext(crop_file)[0]  # ex: eagle.png -> eagle

        draw_header(f"Crop {i}/{len(crops)}: {crop_file}  (GT crop-level: {crop_gt})")
        draw_images_row(crop_file)

        # Predictions table
        y = table_top
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margin_x, y, "Predictions + status (baseline / crop_black / crop_white)")
        y -= 18

        c.setFont("Helvetica", 9)
        for model in models:
            base = row_baseline(model)
            blk = row_black(model, crop_file)
            wht = row_white(model, crop_file)

            b_status, b_pred, b_fp, b_nfp = format_pred_block(crop_gt, base)
            k_status, k_pred, k_fp, k_nfp = format_pred_block(crop_gt, blk)
            w_status, w_pred, w_fp, w_nfp = format_pred_block(crop_gt, wht)

            c.setFont("Helvetica-Bold", 9)
            c.drawString(margin_x, y, f"- {model}")
            y -= line_h

            c.setFont("Helvetica", 9)
            y = write_wrapped(c, f"baseline: {b_status}  pred=[{b_pred}]  fp=[{b_fp}]  n_fp={b_nfp}", margin_x + 14, y, 160, line_h)
            y = write_wrapped(c, f"crop_black: {k_status}  pred=[{k_pred}]  fp=[{k_fp}]  n_fp={k_nfp}", margin_x + 14, y, 160, line_h)
            y = write_wrapped(c, f"crop_white: {w_status}  pred=[{w_pred}]  fp=[{w_fp}]  n_fp={w_nfp}", margin_x + 14, y, 160, line_h)
            y -= 6

            # evita estourar página; se estiver no fim, nova página (continua mesmo crop)
            if y < 60:
                c.showPage()
                draw_header(f"(cont.) Crop {i}/{len(crops)}: {crop_file}  (GT: {crop_gt})")
                y = page_h - 110
                c.setFont("Helvetica", 9)

        c.showPage()

    c.save()

# -----------------------------
# CLI
# -----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--splits", required=True, help="Ex: test_open ou test_closed (pode ser lista separada por vírgula)")
    ap.add_argument("--image_ids", default="", help="Lista separada por vírgula. Se vazio, pega todos do XLSX para o split.")
    ap.add_argument("--outdir", required=True, help="Diretório de saída dos PDFs")
    ap.add_argument("--tables_xlsx", default="DL_all_vlms_baseline_black_white_tables.xlsx", help="Planilha com aba per_crop")
    ap.add_argument("--models", default="gemma3,llama3_2_vision,qwen2_5_vl", help="Lista separada por vírgula")
    args = ap.parse_args()

    splits = [s.strip() for s in args.splits.split(",") if s.strip()]
    models = [m.strip() for m in args.models.split(",") if m.strip()]

    if not os.path.exists(args.tables_xlsx):
        raise SystemExit(f"[ERRO] Não achei a planilha: {args.tables_xlsx}")

    per_crop = load_per_crop_xlsx(args.tables_xlsx)

    # Se image_ids vazio, coletamos todos da planilha para cada split
    if args.image_ids.strip():
        image_ids = [x.strip() for x in args.image_ids.split(",") if x.strip()]
        by_split = {s: image_ids for s in splits}
    else:
        by_split = {}
        for s in splits:
            ids = sorted({k[2] for k in per_crop.keys() if k[0] == s})
            by_split[s] = ids

    for split in splits:
        ids = by_split.get(split, [])
        if not ids:
            print(f"[WARN] Nenhuma image_id encontrada para split={split}")
            continue

        for image_id in ids:
            out_pdf = os.path.join(args.outdir, split, f"{image_id}.pdf")
            try:
                make_pdf_for_image(out_pdf, split, image_id, per_crop, models)
                print(f"[OK] {out_pdf}")
            except Exception as e:
                print(f"[ERRO] {split}/{image_id}: {e}")

if __name__ == "__main__":
    main()
